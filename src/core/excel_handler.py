import pandas as pd
import logging
from typing import List, Dict, Any, Optional, Tuple
import json
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook

class ExcelHandler:
    """Classe responsável por manipular arquivos Excel."""
    
    def __init__(self) -> None:
        self.logger: logging.Logger = logging.getLogger(__name__)
        self.excel_file: str = "output/pool.xlsx"
        self.date_format: str = "%d/%m/%Y"
        self.green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        self.red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
    def _to_float(value):
        try:
            return float(str(value).replace(',', '.'))
        except (ValueError, TypeError):
            return 0
    
    def read_json_file(self, filepath: str = "output/cards.json") -> Dict[str, Any]:
        """Lê o arquivo JSON com os dados das cartas."""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data: Dict[str, Any] = json.load(f)
            self.logger.info(f"Successfully read JSON file: {filepath}")
            return data
        except Exception as e:
            self.logger.error(f"Error reading JSON file: {str(e)}")
            raise
            
    def _find_empty_column(self, ws) -> Tuple[int, str]:
        """Encontra a primeira coluna vazia após as colunas fixas."""
        # Colunas fixas são 'Nome da Carta' e 'Quantidade'
        fixed_columns = 2
        
        # Começa a verificar após as colunas fixas
        for col in range(fixed_columns + 1, ws.max_column + 2):  # +2 para incluir uma coluna extra se necessário
            # Verifica se a coluna existe
            if col > ws.max_column:
                return col, None
                
            # Verifica se a coluna está vazia
            is_empty = True
            for row in range(2, ws.max_row + 1):  # Começa da linha 2 (após cabeçalho)
                if ws.cell(row=row, column=col).value is not None:
                    is_empty = False
                    break
                    
            if is_empty:
                return col, None
                
        # Se não encontrou coluna vazia, retorna a próxima coluna
        return ws.max_column + 1, None
            
    def _check_date_exists(self, ws, current_date: str) -> bool:
        return False
        """Verifica se a data já existe no Excel."""
        for cell in ws[1]:  # Verifica apenas o cabeçalho
            if cell.value == current_date:
                return True
        return False
            
    def update_excel(self, data: Dict[str, Any]) -> str:
        """Atualiza o arquivo Excel com os novos dados."""
        try:
            # Usa a data do JSON
            json_date: str = data['extraction_date']
            # Converte para o formato dd/mm/yyyy
            date_obj: datetime = datetime.strptime(json_date, "%Y-%m-%d %H:%M:%S")
            current_date: str = date_obj.strftime(self.date_format)
            
            # Carrega o Excel existente ou cria um novo
            if os.path.exists(self.excel_file):
                df: pd.DataFrame = pd.read_excel(self.excel_file)
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                # Verifica se a data já existe
                if self._check_date_exists(ws, current_date):
                    wb.close()
                    self.logger.warning(f"Data {current_date} já foi preenchida anteriormente. Processo interrompido.")
                    return self.excel_file
                    
                empty_col, _ = self._find_empty_column(ws)
                wb.close()
            else:
                df: pd.DataFrame = pd.DataFrame(columns=['Nome da Carta', 'Quantidade'])
                empty_col = 3  # Primeira coluna após as fixas
                
            # Adiciona a nova coluna com a data
            df[current_date] = None
            
            # Atualiza os dados
            for card in data['cards']:
                card_name: str = card['name']
                quantity: int = card['quantity']
                price: float = card['price']
                
                # Procura a carta no DataFrame
                if card_name in df['Nome da Carta'].values:
                    idx: int = df[df['Nome da Carta'] == card_name].index[0]
                    df.at[idx, 'Quantidade'] = quantity
                    df.at[idx, current_date] = price
                else:
                    # Adiciona nova carta
                    new_row: Dict[str, Any] = {
                        'Nome da Carta': card_name,
                        'Quantidade': quantity,
                        current_date: price
                    }
                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
            # Salva o DataFrame no Excel
            df.to_excel(self.excel_file, index=False)
            
            # Aplica as cores baseadas na comparação de valores
            self._apply_colors(current_date)
            
            # Aplica o alinhamento centralizado
            self._apply_alignment()
            
            self.logger.info(f"Excel file updated successfully: {self.excel_file}")
            return self.excel_file
            
        except Exception as e:
            self.logger.error(f"Error updating Excel file: {str(e)}")
            raise
            
    def _apply_colors(self, current_date: str) -> None:
        """Aplica as cores nas células baseado na comparação de valores, mantendo cores anteriores."""
        try:
            # Carrega o arquivo Excel
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Encontra o índice da coluna da data atual
            current_date_col = None
            date_columns = []
            
            # Identifica todas as colunas de data e a coluna atual
            for cell in ws[1]:
                if cell.value == current_date:
                    current_date_col = cell.column
                    date_columns.append((cell.column, cell.value))
                elif isinstance(cell.value, str) and cell.value != 'Nome da Carta' and cell.value != 'Quantidade':
                    date_columns.append((cell.column, cell.value))
            
            if current_date_col is None:
                raise ValueError(f"Column for date {current_date} not found")
            
            # Ordena as colunas de data pela posição (para garantir a ordem cronológica)
            date_columns.sort()
            
            # Para cada coluna de data (exceto a primeira), aplica as cores comparando com a anterior
            for i in range(1, len(date_columns)):
                col_num, col_date = date_columns[i]
                prev_col_num = date_columns[i-1][0]
                
                # Se for a coluna atual ou se a coluna anterior for 'Quantidade', pula
                if col_date == current_date or ws.cell(row=1, column=prev_col_num).value == 'Quantidade':
                    continue
                    
                # Para cada linha (começando da segunda)
                for row in range(2, ws.max_row + 1):
                    current_cell = ws.cell(row=row, column=col_num)
                    previous_cell = ws.cell(row=row, column=prev_col_num)
                    
                    # Se algum dos valores for None, pula
                    if current_cell.value is None or previous_cell.value is None:
                        continue
                    
                    # Compara os valores
                    if current_cell.value > previous_cell.value:
                        current_cell.fill = self.green_fill
                    elif current_cell.value < previous_cell.value:
                        current_cell.fill = self.red_fill
                    # Se for igual, mantém o fundo branco (padrão)
            
            # Agora aplica as cores para a coluna atual (última)
            current_col_num = current_date_col
            prev_col_num = date_columns[-2][0] if len(date_columns) > 1 else None
            
            if prev_col_num:
                for row in range(2, ws.max_row + 1):
                    current_cell = ws.cell(row=row, column=current_col_num)
                    previous_cell = ws.cell(row=row, column=prev_col_num)
                    
                    if current_cell.value is None or previous_cell.value is None:
                        continue
                    
                    if(isinstance(current_cell.value, str)):
                        current_cell.value = current_cell.value.replace(',', '.')
                        current_cell.value = float(current_cell.value)
                  
                    if(isinstance(previous_cell.value, str)):
                        previous_cell.value = previous_cell.value.replace(',', '.')
                        previous_cell.value = float(previous_cell.value)

                    if current_cell.value >  previous_cell.value:
                        current_cell.fill = self.green_fill
                    elif current_cell.value <  previous_cell.value:
                        current_cell.fill = self.red_fill
            
            # Salva as alterações
            wb.save(self.excel_file)
            
        except Exception as e:
            self.logger.error(f"Error applying colors: {str(e)}")
            raise
            """Aplica as cores nas células baseado na comparação de valores."""
            try:
                # Carrega o arquivo Excel
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                # Encontra o índice da coluna da data atual
                current_date_col = None
                for cell in ws[1]:
                    if cell.value == current_date:
                        current_date_col = cell.column
                        break
                
                if current_date_col is None:
                    raise ValueError(f"Column for date {current_date} not found")
                
                # Para cada linha (começando da segunda, pois a primeira é cabeçalho)
                for row in range(2, ws.max_row + 1):
                    current_cell = ws.cell(row=row, column=current_date_col)
                    previous_cell = ws.cell(row=row, column=current_date_col - 1)
                    
                    # Se a célula anterior não tem valor, pula
                    if previous_cell.value is None:
                        continue
                    
                    # Se a célula atual não tem valor, pula
                    if current_cell.value is None:
                        continue
                    
                    # Verifica se a coluna anterior é a coluna de quantidade
                    previous_header = ws.cell(row=1, column=previous_cell.column).value
                    if previous_header == 'Quantidade':
                        continue
                    
                    # Compara os valores
                    if current_cell.value > previous_cell.value:
                        current_cell.fill = self.green_fill
                    elif current_cell.value < previous_cell.value:
                        current_cell.fill = self.red_fill
                    # Se for igual, mantém o fundo branco (padrão)
                
                # Salva as alterações
                wb.save(self.excel_file)
                
            except Exception as e:
                self.logger.error(f"Error applying colors: {str(e)}")
                raise
            
    def _apply_alignment(self) -> None:
        """Aplica o alinhamento centralizado em todas as colunas, exceto 'Nome da Carta'."""
        try:
            # Carrega o arquivo Excel
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Para cada coluna
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                
                # Pula a coluna 'Nome da Carta'
                if header == 'Nome da Carta':
                    continue
                
                # Aplica alinhamento centralizado em todas as células da coluna
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = self.center_alignment
            
            # Salva as alterações
            wb.save(self.excel_file)
            
        except Exception as e:
            self.logger.error(f"Error applying alignment: {str(e)}")
            raise 
